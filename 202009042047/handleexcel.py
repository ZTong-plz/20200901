import csv
from openpyxl import *
from openpyxl.styles import Font,Alignment
import os
import builtins


class HandlerExcel(object):
    # 用户文件位置，可修改
    # csv_file = r'C:/Users/吴胤澎/Desktop/Web Of Science Documents.csv'
    # ex_file = r'C:/Users/吴胤澎/Desktop/Web Of Science Documents.xlsx'
    csv_file = r'D:/Web Of Science Documents.csv'
    ex_file = r'D:/Web Of Science Documents.xlsx'

    wb = None
    ws = None
    artNum = None

    def __init__(self):
        if os.path.exists(self.ex_file):
            os.remove(self.ex_file)
        self.preHandleCSV()
        self.preHandleExcel()
    # def toexcel(self):
    #      if os.path.exists(self.ex_file):
    #         os.remove(self.ex_file)
    #      self.preHandleCSV()
    #      self.preHandleExcel()

    def preHandleCSV(self):
        """处理CSV文件，进行必要数据提取"""
        csv_open = builtins.open(self.csv_file)
        csv_reader = csv.reader(csv_open)
        csv_data = list(csv_reader)
        # 用csv必要数据直接创建excel表格
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Report'
        self.ws.cell(1,1).value = "Accession Number"
        self.ws.cell(1,2).value = "Article Title"
        count = 0
        try:
            for i in range(1, len(csv_data)):
                # print(csv_data[i][0] + "," + csv_data[i][3], end='\n')
                self.ws.cell(count+2,2).value = csv_data[i][3]
                self.ws.cell(count+2,1).value = csv_data[i][0]
                count += 1
        except Exception:
            pass
        self.artNum = count
        self.wb.save(self.ex_file)

    def preHandleExcel(self):
        """插入必要字段，调整字体、列宽"""

        self.ws.insert_cols(2)  # 插入‘日期’、‘被引频次’列
        self.ws['b1'] = '日期'
        self.ws.insert_cols(2)
        self.ws['b1'] = '被引频次'

        # 调整列宽
        self.ws.column_dimensions['A'].width = 20  # wos
        self.ws.column_dimensions['B'].width = 20  # 被引频次
        self.ws.column_dimensions['C'].width = 20  # 日期
        self.ws.column_dimensions['D'].width = 70  # 论文标题

        # 设置字体
        font = Font(name='等线', bold=False)
        for i in range(self.ws.max_row):
            for j in range(self.ws.max_column):
                self.ws.cell(i+1,j+1).font = font
        for i in range(self.ws.max_row):
            self.ws.cell(i+1,2).alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(i+1,3).alignment = Alignment(horizontal='center', vertical='center')

        self.wb.save(self.ex_file)

    def getWOSList(self):  # 返回WOS列表
        wosList = []
        for i in range(2, self.artNum+2):
            wosList.append(self.ws.cell(row=i, column=1).value.strip("WOS:"))
        return wosList

    @staticmethod
    def getArticleNum(self):  # 获取文章数
        return count

    def writeInByList(self,listForDict):
        """ListForDict，字典列表，批量将字典拆包写入excel"""
        # 字典格式如下：
        # {'name': 'Fuzzy best-worst multi-criteria decision-making method and its applications', 'cites': '77', 'date': 'APR 1 2017'}
        for element in listForDict:
            wosValue = element['wos']
            for i in range(2, self.artNum + 1):
                if self.ws.cell(i, 1).value == wosValue:
                    self.ws.cell(i, 2).value = element['cites']
                    self.ws.cell(i, 3).value = element['date']
        self.wb.save(self.ex_file)

    def writeInBySingle(self,dic):
        """Dict，单条记录字典，逐条写入EXCEL表中"""
        # 如果用pandas的dataframe方法会改变excel表格的格式，不方便，于是直接用openpyxl提供的api操作
        '''data = pd.read_excel(self.ex_file, index_col='Article Title')  # 论文标题
        nameValue = dic['name']
        data.loc[nameValue, '被引频次'] = dic['cites']
        data.loc[nameValue, ' 日期'] = dic['date']
        DataFrame(data).to_excel(self.ex_file, index=False, header=True)'''
        # nameValue = dic['name']
        wosValue = dic['wos']
        for i in range(2,self.artNum+2):
            if self.ws.cell(i,1).value == wosValue:
                self.ws.cell(i,2).value = dic['cites']
                self.ws.cell(i,3).value = dic['date']
        self.wb.save(self.ex_file)


if __name__ == '__main__':
    test = HandlerExcel()
    DICT = {'wos': 'WOS:000313761800034', 'cites': '77', 'date': 'APR 1 2017'}
    test.writeInBySingle(DICT)
    li = test.getWOSList()
    count = 0
    for i in li:
        count += 1
        print(count,end=':')
        print(i)